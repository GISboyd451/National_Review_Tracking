
import numpy as np
import pandas as pd 
import os
import sys
import datetime
import calendar
from glob import glob
from shutil import copyfile

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side
##
#
#### Globals ####
qc_output_root = r'\\blm\\dfs\\loc\\EGIS\\ProjectsNational\\NationalDataQuality\\Sprint\\analysis_tools\\National_Review'
qc_reports_root = r'\\blm\\dfs\\loc\\EGIS\\ProjectsNational\\NationalDataQuality\\Sprint\\Reports'
backup = r'\\blm\\dfs\\loc\\EGIS\\ProjectsNational\\NationalDataQuality\\Sprint\\analysis_tools\\National_Review\\raw\\Nat_review_backups' #location of backups
national_xlsx = r'\\blm\\dfs\\loc\\EGIS\\ProjectsNational\\NationalDataQuality\\Sprint\\analysis_tools\\National_Review\\National_Review_Tracking.xlsx'
##date_of_qc_run = '20200114' # Format: YYYYMMDD ; File Folder will be something like: 20200114_Reports ;
if (sys.version_info > (3, 0)):
    # Python 3 code in this block
    a = 1
    print ('Date formate: YYYYMMDD, Example: 20200204')
    date_of_qc_run = str(input('Date of QC run": '))
else:
    # Python 2 code in this block
    a = 2
    print ('Date formate: YYYYMMDD, Example: 20200204')
    date_of_qc_run = str(raw_input('Date of QC run": '))

reports_dir = qc_reports_root+os.sep+'%s_Reports' % date_of_qc_run
# Get list of all  files ; Remove .xlsx and .zip
file_list = os.listdir(reports_dir)
file_list = [ x for x in file_list if ".xlsx" not in x ]
file_list = [ x for x in file_list if ".zip" not in x ]
# Present #
present_date = datetime.date.today()
month_name = calendar.month_abbr[present_date.month]
year = present_date.year
month_year = str(month_name) + ' '+ str(year)
#
# Past #
now = datetime.datetime.now()
last_month = now.month-1 if now.month > 1 else 12
if str(month_name) == 'Jan':
    last_year = now.year - 1 # To account for the jan to december subtraction based on current month
else:
    last_year = now.year
prev_month_name = calendar.month_abbr[last_month]
past_month_year = str(prev_month_name)+ ' '+ str(last_year)
#
#### End Globals ####

###### National Tasking Sheet ######
try:
    xlsx = pd.ExcelFile(national_xlsx)
    xlsx = pd.read_excel(xlsx)
    # Create Backup
    if os.path.exists(backup+os.sep+'Nat_Review_%s.xlsx' % past_month_year):
        print ("National Tracking File found.")
        print ('Backup Exists, passing.....')
        pass
    else:
        copyfile(national_xlsx, backup+os.sep+'Nat_Review_%s.xlsx' % past_month_year)
        print ("National Tracking File found. Backup Created.")
except:
    print ("National Tracking File not found.")
####### National Tasking Sheet #####
columns = list(xlsx.columns.values)

#### Start Compile ####
r = pd.DataFrame() # Empty dataframe to append to

for f in file_list:
    state = f.split('_')[0]
    report = pd.ExcelFile(reports_dir + os.sep + f + os.sep + '%s_Quality_Reports.xlsx' % state)
    tab_names = report.sheet_names # All sheet names from report
    
    for s in tab_names:
        report = pd.ExcelFile(reports_dir + os.sep + f + os.sep + '%s_Quality_Reports.xlsx' % state)
        # Read single sheet and use sheet_name vs sheetname based on python ver
        if a == 1:
            report = pd.read_excel(report, skiprows=5, sheet_name=s)
            # Drop Completely empty rows
            report.dropna(axis=0, how='any', inplace=True)
            report.columns = [month_year, month_year+'.1', month_year+'.2', 'Info3', 'Info4']
            report['Info1'] = state # Assign state column
            report[month_year+'.3'] = 0
            r = r.append(report,ignore_index=True)
        else:
            report = pd.read_excel(report, skiprows=5, sheetname=s)
            # Drop Completely empty rows
            report.dropna(axis=0, how='any', inplace=True)
            report.columns = [month_year, month_year+'.1', month_year+'.2', 'Info3', 'Info4']
            report['Info1'] = state # Assign state column
            report[month_year+'.3'] = 0
            r = r.append(report,ignore_index=True)
print ('Finished Compile.')


df = pd.merge(xlsx, r, on=['Info1','Info3','Info4'], how="left") #indicator=True
new_cols = [month_year,month_year+'.1',month_year+'.2',month_year+'.3']
# Perform Percent Change calculation for new month
#df_columns = list(df.columns.values)
df[month_year+'.3'] = df[month_year+'.2'] - df[past_month_year+'.2']
#
row2add = ['Pass Count', 'Total Count', 'Accuracy (%)', 'Percent Change']
df.loc[0, new_cols[0]] = row2add[0]
df.loc[0, new_cols[1]] = row2add[1]
df.loc[0, new_cols[2]] = row2add[2]
df.loc[0, new_cols[3]] = row2add[3]

# Remove NaN, NA from df
df = df.fillna("")
print ('Finished Calcs & Merge.')

df.to_excel(national_xlsx, index=False)

print ('Formatting.....')
#### Formatting ####
file = national_xlsx
wb = load_workbook(file)
ws = wb['Sheet1']
#### set formatting objects ####

## Format top row ##
row1_unbold = Font(bold=False)
row1_nobdr = Border(left=Side(border_style=None, color=None), right=Side(border_style=None, color=None))
for cell in ws["1:1"]:
    cell.font = row1_unbold
    cell.border = row1_nobdr

## Info4 (column D) Bold right border ##
right_border_D = Border(right=Side(style='thick'))
for cell in ws["D:D"]:
    cell.border = right_border_D
## Column D Bold right border End ##

## Row 2 bold border & Bold text ##
TOP_BOTTOM_row2 = Border(top=Side(style='thick'), bottom=Side(style='thick')) 
row_2_bold = Font(bold=True)
for cell in ws["2:2"]:
    cell.font = row_2_bold
    cell.border = TOP_BOTTOM_row2
## Row 2 bold border End ##

## Info4 (column D) Bold right border ##
right_border_D = Border(right=Side(style='thick'))
D2_cell_format = Border(right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
for Dcell in ws["D:D"]:
    if Dcell == ws["D2"]:
        Dcell.border = D2_cell_format
    else:
        Dcell.border = right_border_D
## Column D Bold right border End ##

wb.save(filename=file)
####### Adjust column widths #######
wb = openpyxl.load_workbook(filename = file)        
ws = wb.active

for column_cells in ws.columns:
    length = max(len(str(cell.value)) for cell in column_cells)
    ws.column_dimensions[column_cells[0].column_letter].width = length

wb.save(filename=file)
########## End Adjust ############
print ('End Script.')
